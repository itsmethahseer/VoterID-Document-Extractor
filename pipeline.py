from fastapi import FastAPI, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import cv2
import numpy as np
import pytesseract
from pytesseract import Output
from pdf2image import convert_from_path
from openpyxl import Workbook
from pathlib import Path
import tempfile
from docx import Document  # New import for DOCX

# Create FastAPI instance
app = FastAPI()

# Allow CORS (Cross-Origin Resource Sharing) if needed
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # You can specify the origins if needed, e.g., ["http://localhost:3000"]
    allow_credentials=True,
    allow_methods=["*"],  # Or specify allowed methods like ["GET"]
    allow_headers=["*"],  # Or specify allowed headers
)

# Path for saving output files
OUTPUT_DIR = Path("data/outputs2")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# OCR Configuration
pytesseract.pytesseract.tesseract_cmd = r'/usr/bin/tesseract'
OCR_CONFIG = r'--psm 6 -l hin'  # Hindi language

def preprocess_image(image):
    """Preprocess image for grid detection and OCR."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 128, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    return thresh

def detect_grid(image):
    """Detect grid and extract individual cells."""
    # Detect horizontal and vertical lines
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
    horizontal_lines = cv2.morphologyEx(image, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
    vertical_lines = cv2.morphologyEx(image, cv2.MORPH_OPEN, vertical_kernel, iterations=2)

    # Combine lines to detect grid
    grid = cv2.add(horizontal_lines, vertical_lines)

    # Find contours to extract cells
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cell_boxes = [cv2.boundingRect(c) for c in contours]

    # Sort cells top-to-bottom, left-to-right
    cell_boxes = sorted(cell_boxes, key=lambda b: (b[1], b[0]))
    return cell_boxes

def extract_text_from_cell(image, cell_box):
    """Extract text from a specific cell using OCR."""
    x, y, w, h = cell_box
    cell_image = image[y:y+h, x:x+w]
    text = pytesseract.image_to_string(cell_image, config=OCR_CONFIG)
    return text.strip()

def create_excel_from_grid(grid_data, output_path):
    """Create an Excel file from the extracted grid data."""
    wb = Workbook()
    ws = wb.active

    # Write the grid data to the Excel sheet
    for i, row in enumerate(grid_data):
        for j, cell_text in enumerate(row):
            ws.cell(row=i+1, column=j+1, value=cell_text)

    # Save the workbook to the specified path
    wb.save(output_path)

def create_docx_from_grid(grid_data, output_path):
    """Create a DOCX file from the extracted grid data."""
    doc = Document()
    table = doc.add_table(rows=len(grid_data), cols=len(grid_data[0]))

    # Write the grid data to the DOCX table
    for i, row in enumerate(grid_data):
        for j, cell_text in enumerate(row):
            table.cell(i, j).text = cell_text

    # Save the DOCX file to the specified path
    doc.save(output_path)

def create_text_from_grid(grid_data, output_path):
    """Create a plain text file from the extracted grid data."""
    with open(output_path, "w", encoding="utf-8") as f:
        for row in grid_data:
            f.write("\t".join(row) + "\n")

@app.post("/extract-text/")
async def extract_text(
    file: UploadFile,
    page_start: int = Form(...),
    page_end: int = Form(...),
    output_format: str = Form(..., description="Choose output format: EXCEL, DOCX, or TEXT"),
):
    """Endpoint to extract Hindi text from PDF and preserve layout."""
    if output_format.lower() not in ["excel", "docx", "text"]:
        raise HTTPException(status_code=400, detail="Invalid output format. Choose 'excel', 'docx', or 'text'")

    # Save uploaded file to a temporary location
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(await file.read())
        temp_pdf_path = tmp_file.name

    try:
        # Convert PDF to images
        pdf_images = convert_from_path(temp_pdf_path, dpi=300)
        if page_start < 1 or page_end > len(pdf_images):
            raise HTTPException(status_code=400, detail="Invalid page range specified.")

        grid_data = []

        # Process specified page range
        for page_number in range(page_start - 1, page_end):
            image = np.array(pdf_images[page_number])
            processed_image = preprocess_image(image)
            cell_boxes = detect_grid(processed_image)
            
            # Extract text from each cell
            page_grid = []
            for box in cell_boxes:
                text = extract_text_from_cell(image, box)
                page_grid.append(text)
            grid_data.append(page_grid)

        # Output files
        file_paths = []

        if "excel" in output_format.lower():
            excel_path = OUTPUT_DIR / f"{file.filename.split('.')[0]}_output.xlsx"
            create_excel_from_grid(grid_data, excel_path)
            file_paths.append(f"http://localhost:8000/download/{excel_path.name}")

        if "docx" in output_format.lower():
            docx_path = OUTPUT_DIR / f"{file.filename.split('.')[0]}_output.docx"
            create_docx_from_grid(grid_data, docx_path)
            file_paths.append(f"http://localhost:8000/download/{docx_path.name}")

        if "text" in output_format.lower():
            text_path = OUTPUT_DIR / f"{file.filename.split('.')[0]}_output.txt"
            create_text_from_grid(grid_data, text_path)
            file_paths.append(f"http://localhost:8000/download/{text_path.name}")

        return {"file_paths": file_paths}
    finally:
        # Ensure the temporary file is deleted
        Path(temp_pdf_path).unlink(missing_ok=True)

@app.get("/download/{file_name}")
async def download_file(file_name: str):
    """Endpoint to download the generated Excel, DOCX, or Text file."""
    file_path = OUTPUT_DIR / file_name
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(file_path)
